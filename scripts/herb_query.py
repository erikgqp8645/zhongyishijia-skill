#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按中药查方剂 — 查询一味中药的历代本草论述及含此药的所有方剂

用法:
  python scripts/herb_query.py <中药名>
  python scripts/herb_query.py 细辛
  python scripts/herb_query.py 麻黄 --sqlite C:/path/to/20120413mssql.sqlite
  python scripts/herb_query.py 桂枝 --max-formulas 30
  python scripts/herb_query.py <中药名> --excel output.xlsx

输出格式: 双段 Markdown
  第一段：本药历代本草论述（按朝代排序）
  第二段：含此药的方剂列表（按朝代排序）
"""

from __future__ import annotations

import argparse
import io
import re
import sqlite3
import sys
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _sqlite_utils import find_sqlite_path, setup_windows_stdout
from _source_map import DYNASTY_ORDER, identify_source_string
from _text_utils import esc

setup_windows_stdout()


def parse_sources(source_str: str) -> list[str]:
    """从多出处字符串中解析出独立出处列表"""
    # 去掉前缀标记（"1.出自" "1." 等）
    text = re.sub(r"^\d+[\.、]", "", source_str)
    # 按 "。/" 或独立章节分割
    parts = re.split(r"(?:[②③④⑤⑥⑦⑧⑨⑩]+\s*出自|[②③④⑤⑥⑦⑧⑨⑩]+\s*《)", text)
    sources = []
    for p in parts:
        p = p.strip()
        if p and len(p) > 2:
            sources.append(p)
    # 备用：直接找书名号内容
    if not sources:
        books = re.findall(r"《([^》]+)》", source_str)
        sources = [f"《{b}》" for b in books if b]
    return sources


def query_herb(
    herb: str,
    sqlite_path: Path,
    max_herb_rows: int = 20,
    max_formula_rows: int = 100,
) -> tuple[list[dict], list[dict]]:
    """查询一味中药的本草论述和含此药的方剂"""
    conn = sqlite3.connect(str(sqlite_path))
    cur = conn.cursor()

    herb_rows: list[dict] = []
    formula_rows: list[dict] = []

    # ── 第 1 段：本药历代本草论述（TypeID=40 = 单味药）─────────────
    cur.execute(
        "SELECT MingCheng, ChuChu, LaiYuan, GongNengZZ, XingWei, GuiJing, FuFang "
        "FROM zysjyj WHERE TypeID=40 AND MingCheng=?",
        (herb,),
    )
    for r in cur.fetchall():
        # r: (MingCheng, ChuChu, LaiYuan, GongNengZZ, XingWei, GuiJing, FuFang)
        source_str = r[1] or ""        # ChuChu
        indication = r[3] or ""         # GongNengZZ
        nature = r[4] or ""             # XingWei
        meridian = r[5] or ""           # GuiJing
        classic_formulas = r[6] or ""   # FuFang

        dyn, book, author = identify_source_string(source_str)

        herb_rows.append({
            "herb": herb,
            "dynasty": dyn,
            "book": book,
            "author": author,
            "source_str": source_str[:200],
            "nature": nature,
            "meridian": meridian,
            "indication": indication,
            "classic_formulas": classic_formulas,
        })

    # ── 第 2 段：含此药的所有方剂（TypeID=39 = 中成药/方剂）────────
    cur.execute(
        "SELECT ID, MingCheng, ChuFang, GongNengZZ, ChuChu "
        "FROM zysjyj WHERE TypeID=39 AND ChuFang LIKE ?",
        (f"%{herb}%",),
    )
    for r in cur.fetchall():
        # r: (ID, MingCheng, ChuFang, GongNengZZ, ChuChu)
        formula_name = r[1] or ""
        prescription = r[2] or ""
        indication = r[3] or ""
        source_str = r[4] or ""

        dyn, book, author = identify_source_string(source_str, extra=prescription)

        formula_rows.append({
            "formula": formula_name,
            "dynasty": dyn,
            "book": book,
            "author": author,
            "prescription": prescription,
            "indication": indication,
        })

    conn.close()

    # 按朝代排序
    herb_rows.sort(key=lambda x: (DYNASTY_ORDER.get(x["dynasty"], 99), x["dynasty"], x["book"]))
    formula_rows.sort(key=lambda x: (DYNASTY_ORDER.get(x["dynasty"], 99), x["dynasty"], x["book"], x["formula"]))

    return herb_rows[:max_herb_rows], formula_rows[:max_formula_rows]



def render_excel(
    herb: str,
    herb_rows: list[dict],
    formula_rows: list[dict],
    output_path: Path,
):
    """将查询结果写入 Excel 文件（多 Sheet）"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 本草论述 ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "本草论述"

    # 表头样式
    hdr_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers1 = ["朝代", "著作", "作者", "性味", "归经", "功能主治"]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for row_idx, row in enumerate(herb_rows, 2):
        values = [
            row.get("dynasty", ""),
            row.get("book", ""),
            row.get("author", ""),
            esc(row.get("nature", ""), 40),
            esc(row.get("meridian", ""), 40),
            esc(row.get("indication", ""), 200),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # 列宽
    col_widths1 = [12, 25, 15, 25, 25, 60]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 含药方剂 ───────────────────────────────────
    ws2 = wb.create_sheet("含药方剂")

    headers2 = ["朝代", "著作", "作者", "方剂名", "处方组成", "主治"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for row_idx, row in enumerate(formula_rows, 2):
        values = [
            row.get("dynasty", "") or "待考",
            esc(row.get("book", ""), 25),
            row.get("author", "") or "",
            esc(row.get("formula", ""), 20),
            esc(row.get("prescription", ""), 80),
            esc(row.get("indication", ""), 80),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    col_widths2 = [12, 25, 15, 22, 60, 60]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 统计总览 ───────────────────────────────────
    ws3 = wb.create_sheet("统计总览")
    stat_font = Font(name="微软雅黑", bold=True, size=11)
    stat_fill = PatternFill("solid", fgColor="D9E1F2")
    stat_align = Alignment(horizontal="left", vertical="center")

    ws3.cell(row=1, column=1, value="查询药材").font = stat_font
    ws3.cell(row=1, column=2, value=herb).alignment = stat_align
    ws3.cell(row=2, column=1, value="本草论述").font = stat_font
    ws3.cell(row=2, column=2, value=f"{len(herb_rows)} 条").alignment = stat_align
    ws3.cell(row=3, column=1, value="含药方剂").font = stat_font
    ws3.cell(row=3, column=2, value=f"{len(formula_rows)} 条").alignment = stat_align
    ws3.cell(row=4, column=1, value="数据来源").font = stat_font
    ws3.cell(row=4, column=2, value="中医世家知识库 20120413mssql.sqlite").alignment = stat_align

    for r in range(1, 5):
        ws3.cell(row=r, column=1).fill = stat_fill
        ws3.cell(row=r, column=2).fill = stat_fill

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 40

    wb.save(output_path)
    print(f"✓ 已保存至: {output_path}", file=sys.stderr)


def render(herb: str, herb_rows: list[dict], formula_rows: list[dict]) -> None:
    """渲染 Markdown 输出"""
    print(f"# 「{herb}」历代本草与方剂论述汇总\n")

    # ── 第 1 段 ──────────────────────────────────────────────
    print(f"## 一、{herb}本药历代本草论述\n")
    if herb_rows:
        print("| 朝代 | 著作 | 作者 | 性味 | 归经 | 功能主治 |")
        print("|:----:|:----:|:----:|:----:|:----:|:--------|")
        for row in herb_rows:
            nature = esc(row["nature"] or "", 40) or "—"
            meridian = esc(row["meridian"] or "", 40) or "—"
            indication = esc(row["indication"] or "", 60) or "—"
            book = esc(row["book"] or "", 30)
            print(f"| {row['dynasty']} | {book} | {row['author']} | {nature} | {meridian} | {indication} |")

        # 经典含方节选
        cf = herb_rows[0].get("classic_formulas", "")
        if cf and len(cf) > 5:
            print(f"\n**经典含方**（节选）：{cf[:200]}\n")
    else:
        print("未找到本草记录。\n")

    # ── 第 2 段 ──────────────────────────────────────────────
    total = len(formula_rows)
    print(f"## 二、含「{herb}」的方剂（共 {total} 条，按朝代排序）\n")
    if formula_rows:
        print("| 朝代 | 著作 | 作者 | 方剂名 | 处方组成（节选） | 主治（节选） |")
        print("|:----:|:----:|:----:|:----:|:--------|:--------|")
        for row in formula_rows:
            prescription = esc(row["prescription"] or "", 50) or "—"
            indication = esc(row["indication"] or "", 50) or "—"
            formula = esc(row["formula"] or "", 20)
            book = esc(row["book"] or "", 25)
            dynasty = row["dynasty"] or "待考"
            author = row["author"] or ""
            print(f"| {dynasty} | {book} | {author} | {formula} | {prescription} | {indication} |")
    else:
        print("未找到含此药的方剂。\n")

    print()
    print("---")
    print(f"*数据来源：中医世家知识库 20120413mssql.sqlite（原始 SQLite）*")
    print(f"*查询药材：{herb}*")
    print(f"*本草论述：{len(herb_rows)} 条 | 含药方剂：{total} 条*")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="按中药查方剂 — 查询一味中药的历代本草论述及含此药的所有方剂",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python scripts/query_herb.py 细辛
  python scripts/query_herb.py 麻黄
  python scripts/query_herb.py 桂枝 --max-formulas 30
  python scripts/query_herb.py 人参 --sqlite C:/path/to/20120413mssql.sqlite
        """,
    )
    parser.add_argument("herb", help="要查询的中药材名称（如：细辛、麻黄、桂枝）")
    parser.add_argument(
        "--sqlite",
        help="SQLite 文件路径（默认自动查找）",
    )
    parser.add_argument(
        "--max-formulas",
        type=int,
        default=100,
        help="最多输出多少条方剂（默认 100）",
    )
    parser.add_argument(
        "--max-herb-rows",
        type=int,
        default=20,
        help="最多输出多少条本草论述（默认 20）",
    )
    parser.add_argument(
        "--excel",
        help="输出为 Excel 文件路径（默认输出 Markdown）",
    )
    args = parser.parse_args()

    try:
        sqlite_path = find_sqlite_path(args.sqlite)
    except FileNotFoundError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"正在查询「{args.herb}」...", file=sys.stderr)
    herb_rows, formula_rows = query_herb(
        args.herb,
        sqlite_path,
        max_herb_rows=args.max_herb_rows,
        max_formula_rows=args.max_formulas,
    )
    print(f"找到 {len(herb_rows)} 条本草记录，{len(formula_rows)} 条方剂记录\n", file=sys.stderr)

    if args.excel:
        render_excel(args.herb, herb_rows, formula_rows, Path(args.excel))
    else:
        render(args.herb, herb_rows, formula_rows)


if __name__ == "__main__":
    main()
